# This file is part of the BTA toolset
# (c) Airbus Group CERT, Airbus Group Innovations and Airbus DS CyberSecurity

import openpyxl

from bta.formatters import Formatter
import StringIO
from collections import defaultdict

from openpyxl.workbook import Workbook
from openpyxl.styles import Style, Color, Font, Border, Side, PatternFill, fills, borders
headcolor = Color(rgb="ff9dc5ff")
oddcolor = Color(rgb="ffffefd4")
evencolor = Color(rgb="ffffd794")
hline = Border(bottom=Side(border_style=borders.BORDER_THICK))
headstyle = Style(font=Font(name='Courier', size=11, bold=True),
                  fill=PatternFill(fill_type=fills.FILL_SOLID,
                                   start_color=headcolor,
                                   end_color=headcolor),
                  border=Border(top=Side(border_style=borders.BORDER_THICK),
                                bottom=Side(border_style=borders.BORDER_THICK)),
                  )
oddstyle =  Style(font=Font(name='Courier', size=11),
                  fill=PatternFill(fill_type=fills.FILL_SOLID,
                                   start_color=oddcolor,
                                   end_color=oddcolor),
              )
evenstyle = Style(font=Font(name='Courier', size=11),
                  fill=PatternFill(fill_type=fills.FILL_SOLID,
                                   start_color=evencolor,
                                   end_color=evencolor),
             )
linestyle = [oddstyle, evenstyle]

@Formatter.register
class Excel(Formatter):
    _name_ = "excel"
    def __init__(self):
        self.wb = Workbook()
        self.wb.properties.title = "BTA report"
        self.wb.properties.creator = "BTA"
        self.reportsheet = self.wb.get_active_sheet()
        self.reportsheet.title = "BTA Report"
        self.reportsheet.show_gridlines = False
        self.indent = 1
        self.sheetnames = set()
    def do_add_table(self, name, lvl, table):
        sht = self.wb.create_sheet()
        shtname = self.reportsheet.bad_title_char_re.sub("", name)
        shtname = shtname[:29]
        sht.title = shtname
        sht.default_column_dimension.bestFit = True
        sht.default_column_dimension.auto_size = True
        self.reportsheet.append({self.indent:'=HYPERLINK("#\'%s\'!A1", "see %s")' % (shtname, name)})

        lengths=defaultdict(int)
        hlines = []
        for rownb,row in enumerate(table):
            if row:
                sht.append([""]*lvl+row)
                for c,val in enumerate(row):
                    lengths[c] = max(lengths[c],len(val))
            else:
                hlines.append(rownb)
        hlines.append(sht.max_row)

        for col in range(1, sht.max_column+1):
            sht.cell(row=1, column=col).style = headstyle
        for row in range(2,sht.max_row+1):
            styl = [oddstyle,evenstyle][row%2]
            for col in range(1,sht.max_column+1):
                cell = sht.cell(row=row, column=col)
                cell.style = styl

        for coldim in sht.column_dimensions.itervalues():
            col = openpyxl.worksheet.column_index_from_string(coldim.index)
            coldim.width = lengths[col-1]

    def add_table(self, name, table):
        self.do_add_table(name, 0, table)

    def add_list(self, name, lvl, lst):
        self.do_add_table(name, lvl, lst)

    def add_section(self, section_name, lvl):
        self.reportsheet.append({lvl+1: section_name})
        self.indent=lvl+2

    def add_content(self, content):
        self.reportsheet.append({self.indent: content})

    def finalize(self, encoding=None):
        s = StringIO.StringIO()
        self.wb.save(s)
        return s.getvalue()
